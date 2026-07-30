# ============================================================
# Target-pair co-occurrence analysis by community engagement
#
# Main purpose:
# 1. Identify which target pairs occur more frequently in
#    high-engagement initiatives than in low-engagement initiatives.
# 2. Conduct the analysis for:
#       - Total engagement
#       - Actor breadth
#       - Participatory power
#       - Institutionalization
# 3. Retain target pairs occurring at least twice.
# 4. Produce detailed Excel tables and heatmaps.
#
# Main recommended indicator:
# Delta_Jaccard = High-engagement Jaccard co-occurrence
#                 - Low-engagement Jaccard co-occurrence
# ============================================================

import re
from pathlib import Path
from itertools import combinations

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt


# ============================================================
# 1. File paths and settings
# ============================================================

# Directory containing this script (the repository root)
PROJECT_DIR = Path(__file__).resolve().parent

# Expected input-data location inside the repository
FILE_PATH = PROJECT_DIR / "data" / "Source Data.xlsx"

SHEET_NAME = "Initiative-level data"

# All generated tables and figures will be saved here
OUTPUT_DIR = PROJECT_DIR / "results"

# Columns containing SDG targets
TARGET_COLUMNS = [
    "SDG11-Sub1",
    "SDG11-Sub2",
    "SDG11-Sub3",
    "SDG-X1",
    "SDG-X2",
]

# Engagement variables
ENGAGEMENT_VARIABLES = {
    "Total engagement": "Engagement_Total",
    "Actor breadth": "Norm_AB",
    "Participatory power": "Norm_PP",
    "Institutionalization": "Norm_IN",
}

# Minimum total number of occurrences required for a pair
MIN_PAIR_TOTAL_COUNT = 2

# Grouping method:
# "median"  = low <= median; high > median
# "tertile" = compare bottom third and top third
SPLIT_METHOD = "median"

# Include external SDGs such as SDG7 and SDG13
INCLUDE_EXTERNAL_SDGS = True

# Also repeat the analysis after removing target 11.3
RUN_EXCLUDING_11_3 = True

# Figure settings
FIGSIZE = (10, 8)
DPI = 300


# ============================================================
# 2. Target-processing functions
# ============================================================

def normalize_target(value):
    """
    Standardize target labels.

    Examples:
        11.1  -> 11.1
        11.10 -> 11.1
        11A   -> 11.a
        11-a  -> 11.a
        SDG 7 -> SDG7
        sdg13 -> SDG13
    """

    if pd.isna(value):
        return None

    text = str(value).strip()

    if text == "":
        return None

    if text.lower() in {"nan", "none", "na", "n/a"}:
        return None

    text = text.replace(" ", "")
    text = text.replace("–", "-").replace("—", "-")

    # Standardize numerical SDG 11 targets
    if re.fullmatch(r"11\.\d+", text):
        suffix = text.split(".", 1)[1]

        # Excel may convert 11.1 to 11.10
        suffix = suffix.rstrip("0") or "0"

        return f"11.{suffix}"

    # Standardize 11.a, 11.b and 11.c
    match_11_letter = re.fullmatch(
        r"11[.\-]?([abcABC])",
        text,
    )

    if match_11_letter:
        return f"11.{match_11_letter.group(1).lower()}"

    # Standardize external SDGs
    match_sdg = re.fullmatch(
        r"(?:SDG|Goal)[.\-]?([0-9]+)",
        text,
        flags=re.I,
    )

    if match_sdg:
        return f"SDG{int(match_sdg.group(1))}"

    return text


def target_sort_key(target):
    """
    Sort SDG 11 targets first, followed by external SDGs.

    Example order:
    11.1, 11.2, ..., 11.7, 11.a, 11.b, SDG1, SDG2, ...
    """

    if target.startswith("11."):
        suffix = target[3:]

        if suffix.isdigit():
            return 0, int(suffix), ""

        return 0, 100, suffix

    match = re.fullmatch(r"SDG(\d+)", target)

    if match:
        return 1, int(match.group(1)), ""

    return 2, 999, target


def extract_target_set(row, exclude_11_3=False):
    """
    Extract a unique target set for one initiative.

    Repeated targets within the same initiative are counted only once.
    """

    targets = set()

    for col in TARGET_COLUMNS:
        if col not in row.index:
            continue

        target = normalize_target(row[col])

        if target is not None:
            targets.add(target)

    if not INCLUDE_EXTERNAL_SDGS:
        targets = {
            target
            for target in targets
            if target.startswith("11.")
        }

    if exclude_11_3:
        targets.discard("11.3")

    return targets


# ============================================================
# 3. Engagement-group functions
# ============================================================

def assign_engagement_groups(series, method="median"):
    """
    Divide initiatives into low- and high-engagement groups.

    Median method:
        Low  = value <= median
        High = value > median

    Tertile method:
        Low  = bottom third
        High = top third
        Middle third is excluded
    """

    numeric = pd.to_numeric(series, errors="coerce")

    groups = pd.Series(
        index=series.index,
        dtype="object",
    )

    if method == "median":

        threshold = numeric.median()

        groups.loc[numeric <= threshold] = "Low"
        groups.loc[numeric > threshold] = "High"

        cut_info = {
            "method": "median",
            "threshold": threshold,
        }

    elif method == "tertile":

        lower_threshold = numeric.quantile(1 / 3)
        upper_threshold = numeric.quantile(2 / 3)

        groups.loc[numeric <= lower_threshold] = "Low"
        groups.loc[numeric >= upper_threshold] = "High"

        cut_info = {
            "method": "tertile",
            "lower_threshold": lower_threshold,
            "upper_threshold": upper_threshold,
        }

    else:
        raise ValueError(
            "SPLIT_METHOD must be either 'median' or 'tertile'."
        )

    return groups, cut_info


# ============================================================
# 4. Co-occurrence functions
# ============================================================

def count_target_pairs(target_sets):
    """
    Count:
    1. Individual target occurrences
    2. Unordered target-pair occurrences

    Example:
        Target set = {11.1, 11.3, SDG13}

    Generated pairs:
        11.1–11.3
        11.1–SDG13
        11.3–SDG13
    """

    pair_counts = {}
    target_counts = {}

    for targets in target_sets:

        sorted_targets = sorted(
            targets,
            key=target_sort_key,
        )

        # Count individual targets
        for target in sorted_targets:
            target_counts[target] = (
                target_counts.get(target, 0) + 1
            )

        # Count all possible target pairs
        for target_a, target_b in combinations(
            sorted_targets,
            2,
        ):
            pair = (target_a, target_b)

            pair_counts[pair] = (
                pair_counts.get(pair, 0) + 1
            )

    return pair_counts, target_counts


def safe_divide(numerator, denominator):
    """Safely divide two values."""

    if denominator == 0:
        return np.nan

    if pd.isna(denominator):
        return np.nan

    return numerator / denominator


# ============================================================
# 5. Main target-pair analysis
# ============================================================

def analyze_one_engagement(
    df,
    engagement_col,
    analysis_name,
    exclude_11_3=False,
):
    """
    Analyze target-pair co-occurrence for one engagement variable.

    Indicators:

    Pair prevalence:
        Number of initiatives containing both targets
        divided by the total number of initiatives in the group.

    Jaccard co-occurrence:
        N(A and B) / N(A or B)

    Conditional probability:
        P(B | A) = N(A and B) / N(A)
        P(A | B) = N(A and B) / N(B)
    """

    work = df.copy()

    # Convert engagement measure to numeric
    work[engagement_col] = pd.to_numeric(
        work[engagement_col],
        errors="coerce",
    )

    # Retain initiatives with a valid engagement value
    work = work.loc[
        work[engagement_col].notna()
    ].copy()

    # Extract target sets
    work["Target_Set"] = work.apply(
        extract_target_set,
        axis=1,
        exclude_11_3=exclude_11_3,
    )

    # At least two targets are needed to form a target pair
    work = work.loc[
        work["Target_Set"].map(len) >= 2
    ].copy()

    # Assign high- and low-engagement groups
    work["Engagement_Group"], cut_info = (
        assign_engagement_groups(
            work[engagement_col],
            SPLIT_METHOD,
        )
    )

    # Remove the middle group when using tertiles
    work = work.loc[
        work["Engagement_Group"].isin(["Low", "High"])
    ].copy()

    group_sizes = (
        work["Engagement_Group"]
        .value_counts()
        .to_dict()
    )

    n_low = int(group_sizes.get("Low", 0))
    n_high = int(group_sizes.get("High", 0))

    # Counts for the entire valid sample
    all_pair_counts, all_target_counts = (
        count_target_pairs(work["Target_Set"])
    )

    # Counts in low-engagement initiatives
    low_target_sets = work.loc[
        work["Engagement_Group"] == "Low",
        "Target_Set",
    ]

    low_pair_counts, low_target_counts = (
        count_target_pairs(low_target_sets)
    )

    # Counts in high-engagement initiatives
    high_target_sets = work.loc[
        work["Engagement_Group"] == "High",
        "Target_Set",
    ]

    high_pair_counts, high_target_counts = (
        count_target_pairs(high_target_sets)
    )

    rows = []

    for pair, total_count in all_pair_counts.items():

        # Retain only pairs occurring at least twice
        if total_count < MIN_PAIR_TOTAL_COUNT:
            continue

        target_a, target_b = pair

        low_count = low_pair_counts.get(pair, 0)
        high_count = high_pair_counts.get(pair, 0)

        # ----------------------------------------------------
        # Pair prevalence among all initiatives in each group
        # ----------------------------------------------------

        low_prevalence = safe_divide(
            low_count,
            n_low,
        )

        high_prevalence = safe_divide(
            high_count,
            n_high,
        )

        # ----------------------------------------------------
        # Jaccard co-occurrence
        #
        # Pair count divided by the number of initiatives
        # containing either target A or target B.
        # ----------------------------------------------------

        low_union = (
            low_target_counts.get(target_a, 0)
            + low_target_counts.get(target_b, 0)
            - low_count
        )

        high_union = (
            high_target_counts.get(target_a, 0)
            + high_target_counts.get(target_b, 0)
            - high_count
        )

        low_jaccard = safe_divide(
            low_count,
            low_union,
        )

        high_jaccard = safe_divide(
            high_count,
            high_union,
        )

        # ----------------------------------------------------
        # Conditional probabilities
        # ----------------------------------------------------

        low_p_b_given_a = safe_divide(
            low_count,
            low_target_counts.get(target_a, 0),
        )

        high_p_b_given_a = safe_divide(
            high_count,
            high_target_counts.get(target_a, 0),
        )

        low_p_a_given_b = safe_divide(
            low_count,
            low_target_counts.get(target_b, 0),
        )

        high_p_a_given_b = safe_divide(
            high_count,
            high_target_counts.get(target_b, 0),
        )

        rows.append({
            "Analysis": analysis_name,

            "Target_A": target_a,
            "Target_B": target_b,

            "Total_pair_count": total_count,
            "Low_pair_count": low_count,
            "High_pair_count": high_count,

            "N_low_pair_initiatives": n_low,
            "N_high_pair_initiatives": n_high,

            "Low_pair_prevalence": low_prevalence,
            "High_pair_prevalence": high_prevalence,

            "Delta_pair_prevalence":
                high_prevalence - low_prevalence,

            "Delta_pair_prevalence_pp":
                100 * (
                    high_prevalence
                    - low_prevalence
                ),

            "Low_Jaccard": low_jaccard,
            "High_Jaccard": high_jaccard,

            "Delta_Jaccard":
                high_jaccard - low_jaccard,

            "Low_P_B_given_A": low_p_b_given_a,
            "High_P_B_given_A": high_p_b_given_a,

            "Delta_P_B_given_A":
                high_p_b_given_a
                - low_p_b_given_a,

            "Low_P_A_given_B": low_p_a_given_b,
            "High_P_A_given_B": high_p_a_given_b,

            "Delta_P_A_given_B":
                high_p_a_given_b
                - low_p_a_given_b,

            "Low_target_A_count":
                low_target_counts.get(target_a, 0),

            "High_target_A_count":
                high_target_counts.get(target_a, 0),

            "Low_target_B_count":
                low_target_counts.get(target_b, 0),

            "High_target_B_count":
                high_target_counts.get(target_b, 0),
        })

    result = pd.DataFrame(rows)

    if not result.empty:

        result = result.sort_values(
            by=[
                "Delta_Jaccard",
                "Delta_pair_prevalence_pp",
                "Total_pair_count",
            ],
            ascending=[
                False,
                False,
                False,
            ],
        ).reset_index(drop=True)

    metadata = {
        "Analysis": analysis_name,
        "Engagement_column": engagement_col,
        "Exclude_11_3": exclude_11_3,
        "Grouping_method": cut_info.get("method"),
        "Threshold": cut_info.get("threshold", np.nan),
        "Lower_threshold":
            cut_info.get("lower_threshold", np.nan),
        "Upper_threshold":
            cut_info.get("upper_threshold", np.nan),
        "Valid_pair_initiatives": len(work),
        "N_low": n_low,
        "N_high": n_high,
        "Minimum_pair_count": MIN_PAIR_TOTAL_COUNT,
    }

    return result, metadata, work


# ============================================================
# 6. Matrix functions
# ============================================================

def make_symmetric_matrix(result, value_col):
    """
    Convert pair-level long-form results into a symmetric matrix.
    """

    if result.empty:
        return pd.DataFrame()

    targets = sorted(
        set(result["Target_A"]).union(
            set(result["Target_B"])
        ),
        key=target_sort_key,
    )

    matrix = pd.DataFrame(
        np.nan,
        index=targets,
        columns=targets,
    )

    for _, row in result.iterrows():

        target_a = row["Target_A"]
        target_b = row["Target_B"]
        value = row[value_col]

        matrix.loc[target_a, target_b] = value
        matrix.loc[target_b, target_a] = value

    np.fill_diagonal(
        matrix.values,
        0,
    )

    return matrix


# ============================================================
# 7. Heatmap functions
# ============================================================

def save_heatmap(
    matrix,
    title,
    output_path,
    value_label,
    annotate=False,
):
    """
    Save a target-pair heatmap using matplotlib.
    """

    if matrix.empty:
        return

    figure_width = max(
        FIGSIZE[0],
        len(matrix.columns) * 0.65,
    )

    figure_height = max(
        FIGSIZE[1],
        len(matrix.index) * 0.55,
    )

    fig, ax = plt.subplots(
        figsize=(figure_width, figure_height)
    )

    values = matrix.values.astype(float)

    # Symmetric scale around zero
    finite_values = values[np.isfinite(values)]

    if len(finite_values) > 0:
        maximum_absolute_value = np.max(
            np.abs(finite_values)
        )
    else:
        maximum_absolute_value = 1

    if maximum_absolute_value == 0:
        maximum_absolute_value = 1

    image = ax.imshow(
        values,
        aspect="auto",
        cmap="RdBu_r",
        vmin=-maximum_absolute_value,
        vmax=maximum_absolute_value,
    )

    ax.set_xticks(
        np.arange(len(matrix.columns))
    )

    ax.set_yticks(
        np.arange(len(matrix.index))
    )

    ax.set_xticklabels(
        matrix.columns,
        rotation=45,
        ha="right",
        fontsize=9,
    )

    ax.set_yticklabels(
        matrix.index,
        fontsize=9,
    )

    ax.set_title(
        title,
        fontsize=12,
        pad=14,
    )

    # Draw white gridlines between cells
    ax.set_xticks(
        np.arange(-0.5, len(matrix.columns), 1),
        minor=True,
    )

    ax.set_yticks(
        np.arange(-0.5, len(matrix.index), 1),
        minor=True,
    )

    ax.grid(
        which="minor",
        linewidth=0.5,
    )

    ax.tick_params(
        which="minor",
        bottom=False,
        left=False,
    )

    # Optional value annotations
    if annotate:

        for row_index in range(len(matrix.index)):
            for col_index in range(len(matrix.columns)):

                value = values[row_index, col_index]

                if np.isfinite(value):
                    ax.text(
                        col_index,
                        row_index,
                        f"{value:.2f}",
                        ha="center",
                        va="center",
                        fontsize=6,
                    )

    colorbar = fig.colorbar(
        image,
        ax=ax,
        fraction=0.046,
        pad=0.04,
    )

    colorbar.set_label(
        value_label,
        fontsize=10,
    )

    fig.tight_layout()

    fig.savefig(
        output_path,
        dpi=DPI,
        bbox_inches="tight",
    )

    plt.close(fig)


# ============================================================
# 8. Excel-formatting function
# ============================================================

def format_excel_workbook(excel_path):
    """
    Apply basic formatting to the output Excel workbook.
    """

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(excel_path)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for worksheet in workbook.worksheets:

        worksheet.freeze_panes = "A2"

        # Format header row
        for cell in worksheet[1]:

            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        # Automatically adjust column width
        for column_cells in worksheet.columns:

            column_letter = get_column_letter(
                column_cells[0].column
            )

            maximum_length = 0

            for cell in column_cells:

                if cell.value is None:
                    continue

                value_length = len(str(cell.value))

                maximum_length = max(
                    maximum_length,
                    value_length,
                )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                maximum_length + 2,
                28,
            )

        # Format decimals
        header_map = {
            cell.column: cell.value
            for cell in worksheet[1]
        }

        for row in worksheet.iter_rows(
            min_row=2
        ):
            for cell in row:

                column_name = header_map.get(
                    cell.column,
                    "",
                )

                if column_name is None:
                    continue

                column_name = str(column_name)

                if (
                    "prevalence" in column_name
                    or "Jaccard" in column_name
                    or "_P_" in column_name
                ):
                    cell.number_format = "0.000"

                if column_name.endswith("_pp"):
                    cell.number_format = "0.0"

    workbook.save(excel_path)


# ============================================================
# 9. Main program
# ============================================================

def main():

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    if not FILE_PATH.exists():
        raise FileNotFoundError(
            "Input data file not found:\n"
            f"{FILE_PATH}\n\n"
            "Place the dataset at 'data/Source Data.xlsx' "
            "before running the analysis."
        )

    print("Reading source data...")

    df = pd.read_excel(
        FILE_PATH,
        sheet_name=SHEET_NAME,
    )

    # Remove extra spaces from column names
    df.columns = [
        str(column).strip()
        for column in df.columns
    ]

    required_columns = (
        TARGET_COLUMNS
        + list(ENGAGEMENT_VARIABLES.values())
    )

    missing_columns = [
        column
        for column in required_columns
        if column not in df.columns
    ]

    if missing_columns:

        print("\nAvailable columns:")
        for column in df.columns:
            print(column)

        raise KeyError(
            "\nThe following required columns are missing:\n"
            + "\n".join(missing_columns)
        )

    print(f"Total initiatives read: {len(df)}")

    # Two analysis modes:
    # 1. All targets
    # 2. Target 11.3 removed
    analysis_modes = [
        ("All targets", False)
    ]

    if RUN_EXCLUDING_11_3:
        analysis_modes.append(
            ("Exclude 11.3", True)
        )

    excel_path = (
        OUTPUT_DIR
        / "Target_pair_engagement_analysis.xlsx"
    )

    metadata_rows = []
    all_top_pairs = []

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
    ) as writer:

        for mode_label, exclude_11_3 in analysis_modes:

            for analysis_label, engagement_col in (
                ENGAGEMENT_VARIABLES.items()
            ):

                analysis_name = (
                    f"{analysis_label} - {mode_label}"
                )

                print(
                    "\nRunning:",
                    analysis_name,
                )

                result, metadata, grouped_data = (
                    analyze_one_engagement(
                        df=df,
                        engagement_col=engagement_col,
                        analysis_name=analysis_name,
                        exclude_11_3=exclude_11_3,
                    )
                )

                metadata_rows.append(metadata)

                # --------------------------------------------
                # Excel sheet-name abbreviations
                # --------------------------------------------

                name_map = {
                    "Total engagement": "Total",
                    "Actor breadth": "AB",
                    "Participatory power": "PP",
                    "Institutionalization": "IN",
                }

                mode_short = (
                    "All"
                    if not exclude_11_3
                    else "No113"
                )

                measure_short = name_map[
                    analysis_label
                ]

                # --------------------------------------------
                # Detailed pair-level results
                # --------------------------------------------

                detail_sheet = (
                    f"Pairs_{measure_short}_{mode_short}"
                )

                result.to_excel(
                    writer,
                    sheet_name=detail_sheet,
                    index=False,
                )

                # --------------------------------------------
                # Jaccard-difference matrix
                # --------------------------------------------

                jaccard_matrix = make_symmetric_matrix(
                    result,
                    "Delta_Jaccard",
                )

                jaccard_sheet = (
                    f"Jaccard_{measure_short}_{mode_short}"
                )

                jaccard_matrix.to_excel(
                    writer,
                    sheet_name=jaccard_sheet,
                )

                # --------------------------------------------
                # Pair-prevalence difference matrix
                # --------------------------------------------

                prevalence_matrix = (
                    make_symmetric_matrix(
                        result,
                        "Delta_pair_prevalence_pp",
                    )
                )

                prevalence_sheet = (
                    f"Prev_{measure_short}_{mode_short}"
                )

                prevalence_matrix.to_excel(
                    writer,
                    sheet_name=prevalence_sheet,
                )

                # --------------------------------------------
                # Conditional probability matrices
                #
                # The matrices are directional:
                # rows = Target A
                # columns = Target B
                # value = change in P(B | A)
                # --------------------------------------------

                if not result.empty:

                    all_targets = sorted(
                        set(result["Target_A"]).union(
                            set(result["Target_B"])
                        ),
                        key=target_sort_key,
                    )

                    conditional_matrix = pd.DataFrame(
                        np.nan,
                        index=all_targets,
                        columns=all_targets,
                    )

                    for _, row in result.iterrows():

                        target_a = row["Target_A"]
                        target_b = row["Target_B"]

                        conditional_matrix.loc[
                            target_a,
                            target_b,
                        ] = row[
                            "Delta_P_B_given_A"
                        ]

                        conditional_matrix.loc[
                            target_b,
                            target_a,
                        ] = row[
                            "Delta_P_A_given_B"
                        ]

                    np.fill_diagonal(
                        conditional_matrix.values,
                        0,
                    )

                else:
                    conditional_matrix = pd.DataFrame()

                conditional_sheet = (
                    f"Cond_{measure_short}_{mode_short}"
                )

                conditional_matrix.to_excel(
                    writer,
                    sheet_name=conditional_sheet,
                )

                # --------------------------------------------
                # Save top-pair summary
                # --------------------------------------------

                if not result.empty:

                    top_result = result.head(15).copy()

                    top_result.insert(
                        0,
                        "Rank",
                        range(1, len(top_result) + 1),
                    )

                    all_top_pairs.append(top_result)

                # --------------------------------------------
                # Save heatmaps
                # --------------------------------------------

                safe_name = re.sub(
                    r"[^A-Za-z0-9_\-]",
                    "_",
                    f"{measure_short}_{mode_short}",
                )

                save_heatmap(
                    matrix=jaccard_matrix,
                    title=(
                        f"Target-pair integration difference\n"
                        f"{analysis_name}"
                    ),
                    output_path=(
                        OUTPUT_DIR
                        / f"Heatmap_Delta_Jaccard_{safe_name}.png"
                    ),
                    value_label=(
                        "High minus low Jaccard co-occurrence"
                    ),
                    annotate=True,
                )

                save_heatmap(
                    matrix=prevalence_matrix,
                    title=(
                        f"Target-pair prevalence difference\n"
                        f"{analysis_name}"
                    ),
                    output_path=(
                        OUTPUT_DIR
                        / f"Heatmap_Delta_Prevalence_{safe_name}.png"
                    ),
                    value_label=(
                        "High minus low prevalence "
                        "(percentage points)"
                    ),
                    annotate=True,
                )

                save_heatmap(
                    matrix=conditional_matrix,
                    title=(
                        f"Conditional target connection difference\n"
                        f"{analysis_name}"
                    ),
                    output_path=(
                        OUTPUT_DIR
                        / f"Heatmap_Delta_Conditional_{safe_name}.png"
                    ),
                    value_label=(
                        "High minus low conditional probability"
                    ),
                    annotate=True,
                )

                # Print the first ten pairs
                if not result.empty:

                    print(
                        result[
                            [
                                "Target_A",
                                "Target_B",
                                "Total_pair_count",
                                "Low_pair_count",
                                "High_pair_count",
                                "Delta_pair_prevalence_pp",
                                "Delta_Jaccard",
                            ]
                        ].head(10)
                    )

                else:
                    print(
                        "No target pairs met the occurrence threshold."
                    )

        # ----------------------------------------------------
        # Metadata sheet
        # ----------------------------------------------------

        metadata_df = pd.DataFrame(
            metadata_rows
        )

        metadata_df.to_excel(
            writer,
            sheet_name="Metadata",
            index=False,
        )

        # ----------------------------------------------------
        # Combined top-pair summary
        # ----------------------------------------------------

        if all_top_pairs:

            combined_top_pairs = pd.concat(
                all_top_pairs,
                ignore_index=True,
            )

            combined_top_pairs.to_excel(
                writer,
                sheet_name="Top_pairs_summary",
                index=False,
            )

    format_excel_workbook(
        excel_path
    )

    print("\nAnalysis completed.")
    print(f"Excel results saved to:\n{excel_path}")
    print(f"\nFigures saved to:\n{OUTPUT_DIR}")


if __name__ == "__main__":
    main()